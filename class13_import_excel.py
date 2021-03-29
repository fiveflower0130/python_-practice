import datetime
from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
# ws['A1'] = 42
datas = [
    ["姓名", "年齡", "身高", "體重"],
    ["Mary", "18", "160", "50"],
    ["John", "20", "175", "65"]
]
# Rows can also be appended
for data in datas:
    ws.append(data)
#ws.append([1, 2, 3, 4])

# Python types will automatically be converted
# ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")
